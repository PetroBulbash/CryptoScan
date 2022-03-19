import copy
import pandas as pd
from openpyxl import load_workbook
from pyunpack import Archive
import os
import shutil
from hashlib import sha256
import docx2txt

digits58 = '123456789ABCDEFGHJKLMNPQRSTUVWXYZabcdefghijkmnopqrstuvwxyz'

pack_formats = ".7z, .ace, .alz, .a, .arc, .arj, .bz2, .cab, .Z, .cpio, .deb, .dms, .gz, .lrz, .lha, .lzh., lz, " \
               ".lzma, .lzo, .rpm, .rar, .rz, .tar, .xz, .zip, .jar, .zoo".split(", ")


# The procedures check_base58check and decode_base58 were initially from:
# http://rosettacode.org/wiki/Bitcoin/address_validation#Python
# released under the GNU Free Documentation License 1.2
def check_base58check(bc, byte_length):
    bcbytes = decode_base58(bc, byte_length)
    if bcbytes == None:
        return False
    return bcbytes[-4:] == sha256(sha256(bcbytes[:-4]).digest()).digest()[:4]


def decode_base58(bc, length):
    n = 0
    # converts the Base58 encoded string to decimal
    for char in bc:
        n = n * 58 + digits58.index(char)
    # At this point this script returns it as bigendian bytes, which is the length which has been set below
    return n.to_bytes(length, 'big')


# Clearing a given directory
def clear_folder(folder):
    for filename in os.listdir(folder):
        file_path = os.path.join(folder, filename)
        try:
            if os.path.isfile(file_path) or os.path.islink(file_path):
                os.unlink(file_path)
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)
        except Exception as e:
            print('Failed to delete %s. Reason: %s' % (file_path, e))


# Get all files in folder and subfolders
def get_folder_content(folder, with_folder_path=True):
    f = []
    for (dirpath, dirnames, filenames) in os.walk(folder):
        f.extend(filenames)
    return f


# Get the necessary information about the file
def file_analitic(file):
    file = file.replace("\\", "/")
    file_name, file_extension = os.path.splitext(file)
    file_info = {"name": file_name, "ext": file_extension, "path": file, "cur_path": file, "in_pack": "", "sh": "", "pack_lev": 0}
    return file_info


# .xls and .xlsx - files transform to txt (each sheet) in tmp-folder
def convert_excel_file(ffile, list_of_files):
    wb = load_workbook(ffile['cur_path'], read_only=True, keep_links=False)
    sheetnames = wb.sheetnames
    folder_out = os.path.dirname(os.path.abspath(__file__)) + "\\tmp\\totxt"
    clear_folder(folder_out)
    for sh in sheetnames:
        f = copy.copy(ffile)
        f["sh"] = sh
        f["ext"] = ""
        f["cur_path"] = folder_out + "\\" + sh
        with open(f["cur_path"], 'w') as file:
            df = pd.read_excel(ffile['cur_path'], sheet_name=sh)
            df.to_string(file, index=False, na_rep='')
            list_of_files.append(f)
    return list_of_files


def convert_docx_file(ffile, list_of_files):
    all_text = docx2txt.process(ffile['cur_path'])
    folder_out = os.path.dirname(os.path.abspath(__file__)) + "\\tmp\\totxt"
    f = copy.copy(ffile)
    f["ext"] = ""
    f["cur_path"] = folder_out + "\\docx_text"
    with open(f["cur_path"], "w") as text_file:
        print(all_text, file=text_file)
    list_of_files.append(f)
    return list_of_files


# packed files unpack in tmp-folder
def convert_pack_file(file, list_of_files):
    folder_out = os.path.dirname(os.path.abspath(__file__)) + "\\tmp\\unpack" + str(file["pack_lev"])
    if not os.path.exists(folder_out):
        os.makedirs(folder_out)
    else:
        clear_folder(folder_out)
    Archive(file['path']).extractall(folder_out)
    files_out_0 = get_folder_content(folder_out)
    for ff in files_out_0:
        f = file_analitic(ff)
        f["pack_lev"] = file["pack_lev"] + 1
        f["path"] = file["path"]
        f["in_pack"] = f["cur_path"]
        f["cur_path"] = folder_out + "\\" + f["cur_path"]
        list_of_files.append(f)
    return list_of_files
